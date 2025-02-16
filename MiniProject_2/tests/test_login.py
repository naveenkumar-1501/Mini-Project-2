"""
test_login.py
Tests login functionality using data from an Excel file.
"""

import pytest
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from MiniProject_2.pages.login_page import LoginPage
from MiniProject_2.common import Config

def load_test_data(file_path):
    """Loads test data from an Excel file."""
    workbook = load_workbook(file_path)
    sheet = workbook.active
    test_data = []
    # Skip the header row and collect data
    for index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True), start=2):test_data.append((index, row[5], row[6]))
    workbook.close()
    return test_data

def write_test_result(file_path, row_num, result):
    """Writes the test result to the Excel file."""
    workbook = load_workbook(file_path)
    sheet = workbook.active
    # Write the result in the 8th column
    sheet.cell(row=row_num, column=8, value=result)
    workbook.save(file_path)
    workbook.close()

@pytest.mark.usefixtures("driver")
class TestLogin:
    @pytest.mark.parametrize("row_num,username,password", load_test_data("test_data.xlsx"))
    def test_login_with_credentials(self, driver, row_num, username, password):
        """Test login functionality."""
        driver.get(Config.URL)
        # Use the LoginPage class
        login_page = LoginPage(driver)
        # Define the error message XPath at the beginning of the method
        error_message_xpath = "//p[text()='Invalid credentials']"
        #Adding cookies before login
        driver.add_cookie({"name": "test_cookie", "value": "cookie_value"})
        print("Cookies before login:", driver.get_cookies())
        try:
            # Call the login method from LoginPage
            login_page.login(username, password)
            # Validate login by checking the dashboard URL or a unique element
            if "dashboard" in driver.current_url.lower():
                print(f"Login successful for user: {username}")
                result = "Pass"
                #Retrieve cookies after login
                print("Cookies after successful login:", driver.get_cookies())
                # Perform logout after successful login
                try:
                    login_page.logout()
                    print(f"Logout successful for user: {username}")
                except Exception as logout_error:
                    print(f"Logout failed for user: {username}. Error: {logout_error}")
                    result = "Fail"
            else:
                # Wait for the error message to appear
                try:
                    error_message = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, error_message_xpath)))
                    if error_message.is_displayed():
                        print(f"Login failed for user: {username} due to invalid credentials.")
                        result = "Fail"
                    else:
                        print(f"Unexpected behavior for user: {username}")
                        result = "Fail"
                except Exception as error:
                    print(f"Error message not found or unexpected behavior: {error}")
                    result = "Fail"
                # Update Excel with the result
                write_test_result("test_data.xlsx", row_num, result)
        except Exception as error:
            # Log and update result as Fail in case of exceptions
            print(f"Exception encountered during login test for user {username}: {error}")
            write_test_result("test_data.xlsx", row_num, "Fail")
            pytest.fail(f"Test failed for username: {username}. Error: {error}")
