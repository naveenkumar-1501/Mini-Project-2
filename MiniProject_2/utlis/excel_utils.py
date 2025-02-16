"""
excel_utils.py
Utility functions for reading and writing Excel files.
"""

import openpyxl
class ExcelUtils:
    @staticmethod
    def read_data(file_path, sheet_name):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)
        return data

    @staticmethod
    def write_data(file_path, sheet_name, row, col, value):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        sheet.cell(row=row, column=col, value=value)
        workbook.save(file_path)
